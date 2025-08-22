#!/usr/bin/env python3
"""
Simple Excel Data Extractor for AIA Lucky Draw System
Uses openpyxl only (no pandas dependency)
"""

import json
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip3 install --user openpyxl")
    sys.exit(1)

def extract_excel_data_simple(excel_path):
    """Extract data from the Excel file using openpyxl only"""
    
    print(f"Reading Excel file: {excel_path}")
    
    try:
        workbook = load_workbook(excel_path, read_only=True)
        print(f"Available sheets: {workbook.sheetnames}")
        
        # Look for value-only sheet first
        value_sheet_names = [sheet for sheet in workbook.sheetnames if 'value' in sheet.lower() and 'only' in sheet.lower()]
        if value_sheet_names:
            print(f"\nFound value-only sheet: {value_sheet_names[0]}")
            generation_ws = workbook[value_sheet_names[0]]
            print(f"Value-only sheet:")
        else:
            # Show Generation sheet structure
            if 'Generation' in workbook.sheetnames:
                generation_ws = workbook['Generation']
                print(f"\nGeneration sheet:")
        
        print(f"Max row: {generation_ws.max_row}, Max col: {generation_ws.max_column}")
        
        # Get headers
        headers = []
        for col in range(1, generation_ws.max_column + 1):
            cell_value = generation_ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Col{col}")
        
        print(f"Headers: {headers}")
        
        # Show first 10 data rows
        print("\nFirst 10 data rows:")
        for row in range(2, min(12, generation_ws.max_row + 1)):
            row_data = []
            for col in range(1, len(headers) + 1):
                cell_value = generation_ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"Row {row}: {row_data}")
        
        # Show Eligible Agent sheet structure
        eligible_sheets = [sheet for sheet in workbook.sheetnames if 'eligible' in sheet.lower() and 'agent' in sheet.lower()]
        if eligible_sheets:
            sheet_name = eligible_sheets[0]
            eligible_ws = workbook[sheet_name]
            print(f"\n{sheet_name} sheet:")
            print(f"Max row: {eligible_ws.max_row}, Max col: {eligible_ws.max_column}")
            
            # Get headers
            headers = []
            for col in range(1, eligible_ws.max_column + 1):
                cell_value = eligible_ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")
            
            print(f"Headers: {headers}")
            
            # Show first 10 data rows
            print("\nFirst 10 data rows:")
            for row in range(2, min(12, eligible_ws.max_row + 1)):
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = eligible_ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"Row {row}: {row_data}")
        
        workbook.close()
        return workbook.sheetnames
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def process_generation_data(excel_path, agent_name_col, group_no_col, lay_see_amount_col):
    """Process Generation sheet data and count prize amounts"""
    
    workbook = load_workbook(excel_path, read_only=True)
    
    # Look for "Value only" or similar sheet first
    value_sheet_names = [sheet for sheet in workbook.sheetnames if 'value' in sheet.lower() and 'only' in sheet.lower()]
    if value_sheet_names:
        generation_ws = workbook[value_sheet_names[0]]
        print(f"Using value-only sheet: {value_sheet_names[0]}")
    else:
        generation_ws = workbook['Generation']
        print("Using Generation sheet (no value-only sheet found)")
    
    # Get column indices (convert from 1-based to 0-based, or use column names)
    headers = []
    for col in range(1, generation_ws.max_column + 1):
        cell_value = generation_ws.cell(row=1, column=col).value
        headers.append(str(cell_value) if cell_value else f"Col{col}")
    
    print(f"Generation headers: {headers}")
    
    # Find column indices
    try:
        agent_name_idx = headers.index(agent_name_col) + 1
        group_no_idx = headers.index(group_no_col) + 1
        lay_see_amount_idx = headers.index(lay_see_amount_col) + 1
    except ValueError as e:
        print(f"Column not found: {e}")
        workbook.close()
        return {}
    
    workers_data = {}
    
    # Process each row
    for row in range(2, generation_ws.max_row + 1):
        agent_name = generation_ws.cell(row=row, column=agent_name_idx).value
        group_no = generation_ws.cell(row=row, column=group_no_idx).value
        lay_see_amount = generation_ws.cell(row=row, column=lay_see_amount_idx).value
        
        if not agent_name or not group_no:
            continue
            
        agent_name = str(agent_name).strip()
        group_no = str(group_no).strip()
        
        # Process lay_see_amount - should be actual values now
        prize_amount = 0
        if lay_see_amount and str(lay_see_amount).strip() not in ['None', '', 'null']:
            try:
                # Try to extract numeric value
                prize_amount = float(str(lay_see_amount).strip())
            except (ValueError, TypeError):
                # If still not a number, skip this entry
                prize_amount = 0
        
        # Create group if not exists
        if group_no not in workers_data:
            workers_data[group_no] = {}
        
        # Create worker if not exists
        if agent_name not in workers_data[group_no]:
            workers_data[group_no][agent_name] = {
                'name': agent_name,
                'group_no': group_no,
                'prize_counts': {},  # Count of each prize amount
                'total_prize_amount': 0,
                'tickets': 0
            }
        
        # Count prize amounts
        worker = workers_data[group_no][agent_name]
        if prize_amount > 0:
            prize_key = f"${int(prize_amount)}" if prize_amount == int(prize_amount) else f"${prize_amount:.2f}"
            if prize_key not in worker['prize_counts']:
                worker['prize_counts'][prize_key] = 0
            worker['prize_counts'][prize_key] += 1
            worker['total_prize_amount'] += prize_amount
        
        worker['tickets'] += 1
    
    workbook.close()
    return workers_data

def process_eligible_agent_data(excel_path, group_no_col, family_col, agent_col, agent_name_col, agency_code_col, district_col):
    """Process Eligible Agent sheet data"""
    
    workbook = load_workbook(excel_path, read_only=True)
    
    # Find eligible agent sheet
    eligible_sheets = [sheet for sheet in workbook.sheetnames if 'eligible' in sheet.lower() and 'agent' in sheet.lower()]
    if not eligible_sheets:
        workbook.close()
        return {}
    
    eligible_ws = workbook[eligible_sheets[0]]
    
    # Get column indices
    headers = []
    for col in range(1, eligible_ws.max_column + 1):
        cell_value = eligible_ws.cell(row=1, column=col).value
        headers.append(str(cell_value) if cell_value else f"Col{col}")
    
    print(f"Eligible Agent headers: {headers}")
    
    # Find column indices
    try:
        group_no_idx = headers.index(group_no_col) + 1
        family_idx = headers.index(family_col) + 1
        agent_idx = headers.index(agent_col) + 1
        agent_name_idx = headers.index(agent_name_col) + 1
        agency_code_idx = headers.index(agency_code_col) + 1
        district_idx = headers.index(district_col) + 1
    except ValueError as e:
        print(f"Column not found: {e}")
        workbook.close()
        return {}
    
    group_families_data = {}
    
    # Process each row
    for row in range(2, eligible_ws.max_row + 1):
        group_no = eligible_ws.cell(row=row, column=group_no_idx).value
        family = eligible_ws.cell(row=row, column=family_idx).value
        agent = eligible_ws.cell(row=row, column=agent_idx).value
        agent_name = eligible_ws.cell(row=row, column=agent_name_idx).value
        agency_code = eligible_ws.cell(row=row, column=agency_code_idx).value
        district = eligible_ws.cell(row=row, column=district_idx).value
        
        if not group_no:
            continue
            
        group_no = str(group_no).strip()
        family = str(family).strip() if family else ""
        agent = str(agent).strip() if agent else ""
        agent_name = str(agent_name).strip() if agent_name else ""
        agency_code = str(agency_code).strip() if agency_code else ""
        district = str(district).strip() if district else ""
        
        # Store group info
        if group_no not in group_families_data:
            group_families_data[group_no] = {
                'family': family,
                'district': district,
                'agents': []
            }
        
        # Update family and district if this row has better info
        if family and not group_families_data[group_no]['family']:
            group_families_data[group_no]['family'] = family
        if district and not group_families_data[group_no]['district']:
            group_families_data[group_no]['district'] = district
        
        # Add agent info
        if agent_name:
            group_families_data[group_no]['agents'].append({
                'agent': agent,
                'agent_name': agent_name,
                'agency_code': agency_code,
                'district': district  # Store individual agent's district
            })
    
    workbook.close()
    return group_families_data

def generate_javascript_data(workers_data, group_families_data):
    """Generate JavaScript data structure"""
    
    js_groups = {}
    group_counter = 1
    
    for group_no, workers in workers_data.items():
        group_id = f"group-{group_counter}"
        
        # Get family and district info
        family_name = group_no  # Default to group number
        district_name = ""
        agents_info = {}
        
        if group_no in group_families_data:
            family_info = group_families_data[group_no]
            family_name = family_info['family'] if family_info['family'] else group_no
            district_name = family_info['district']
            
            # Create a lookup map for agent info
            for agent_info in family_info['agents']:
                agents_info[agent_info['agent_name']] = {
                    'agent': agent_info['agent'],
                    'agency_code': agent_info['agency_code'],
                    'district': agent_info['district']  # Use individual agent's district
                }
        
        # Convert workers
        js_workers = []
        emp_counter = 1
        
        for agent_name, worker_info in workers.items():
            employee_id = f"EMP{group_counter:03d}{emp_counter:03d}"
            
            # Find agency code for this agent
            agent_info = agents_info.get(agent_name, {})
            agent_code = agent_info.get('agent', "")
            agency_code = agent_info.get('agency_code', "")
            agent_district = agent_info.get('district', "")  # Use individual agent's district
            
            js_workers.append({
                'name': agent_name,
                'tickets': worker_info['tickets'],
                'employeeId': employee_id,
                'groupNo': group_no,
                'agent': agent_code,
                'agencyCode': agency_code,
                'district': agent_district,  # Use agent's individual district
                'prizeCounts': worker_info['prize_counts'],
                'totalPrizeAmount': worker_info['total_prize_amount']
            })
            emp_counter += 1
        
        # Create group data
        js_groups[group_id] = {
            'id': group_id,
            'name': family_name,
            'groupNo': group_no,
            'district': district_name,
            'icon': '👨‍👩‍👧‍👦',
            'description': f"{district_name} District" if district_name else f"Group {group_no}",
            'workers': js_workers
        }
        
        group_counter += 1
    
    return js_groups

def main():
    excel_path = "/Users/user/html/aia.luckdraw/aia-lucky-draw-wheel/assets/info/20250811 Lucky Money for Special Districts_Final.xlsm"
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return
    
    # Extract and show data structure first
    sheet_names = extract_excel_data_simple(excel_path)
    
    if sheet_names is None:
        print("Failed to read Excel file")
        return
    
    print("\n" + "="*50)
    print("EXCEL DATA EXTRACTION COMPLETE")
    print("="*50)
    print("\nTo extract data, run with column names:")
    print("python3 simple_extract.py 'Agent Name' 'Group No.' 'Lay See amount' 'Group No.' 'Family' 'Agent' 'Agent name' 'Agency code' 'District'")

if __name__ == "__main__":
    if len(sys.argv) == 10:
        excel_path = "/Users/user/html/aia.luckdraw/aia-lucky-draw-wheel/assets/info/20250811 Lucky Money for Special Districts_Final.xlsm"
        
        # Generation sheet columns
        agent_name_col = sys.argv[1]
        group_no_col = sys.argv[2] 
        lay_see_amount_col = sys.argv[3]
        
        # Eligible Agent sheet columns
        eligible_group_no_col = sys.argv[4]
        family_col = sys.argv[5]
        agent_col = sys.argv[6]
        eligible_agent_name_col = sys.argv[7]
        agency_code_col = sys.argv[8]
        district_col = sys.argv[9]
        
        workers_data = process_generation_data(excel_path, agent_name_col, group_no_col, lay_see_amount_col)
        group_families_data = process_eligible_agent_data(excel_path, eligible_group_no_col, family_col, agent_col, eligible_agent_name_col, agency_code_col, district_col)
        js_data = generate_javascript_data(workers_data, group_families_data)
        
        # Output the JavaScript data
        print("\n" + "="*50)
        print("GENERATED JAVASCRIPT DATA")
        print("="*50)
        print(json.dumps(js_data, indent=2, ensure_ascii=False))
        
        # Save to file
        output_path = "/Users/user/html/aia.luckdraw/aia-lucky-draw-wheel/extracted_data.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(js_data, f, indent=2, ensure_ascii=False)
        print(f"\nData saved to: {output_path}")
    else:
        main()
